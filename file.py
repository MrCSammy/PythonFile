import openpyxl

inv_file = openpyxl.load_workbook("PythonBook.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier = {}
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 2).value
    inventory = product_list.cell(product_row, 3).value
    price = product_list.cell(product_row, 4).value

# calculation for number of similar names
    if supplier_name in product_per_supplier:
        current_no_product = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_no_product + 1
    else:
        product_per_supplier[supplier_name] = 1

        # calculation of total price and amount paid

    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
